"""Tests d'integration des endpoints comptes/gestion (comptes.py).

Chaque test tourne sur une base SQLite EN MEMOIRE (get_db surcharge) : rien
n'est ecrit dans backend/data/. Couvre l'auth enseignant, la gestion de
classe/eleves, le cloisonnement entre enseignants, et la connexion eleve.
"""

from __future__ import annotations

import unittest
from unittest.mock import patch

from fastapi.testclient import TestClient
from sqlalchemy.orm import Session, sessionmaker

import comptes
from database import Progression, create_db_engine, init_db
from main import app


class ComptesIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.engine = create_db_engine("sqlite://")  # en memoire (StaticPool)
        init_db(self.engine)
        testing_session = sessionmaker(
            bind=self.engine, autoflush=False, expire_on_commit=False, class_=Session
        )

        def override_get_db():
            db = testing_session()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[comptes.get_db] = override_get_db
        comptes._reset_tokens()
        self.client = TestClient(app)

    def tearDown(self) -> None:
        app.dependency_overrides.clear()
        comptes._reset_tokens()
        self.engine.dispose()

    # --- Helpers ---
    def _inscrire(self, identifiant="prof1", mot_de_passe="secret123", nom="Prof Un"):
        return self.client.post(
            "/enseignant/inscription",
            json={"nom": nom, "identifiant": identifiant, "mot_de_passe": mot_de_passe},
        )

    def _token(self, identifiant="prof1", mot_de_passe="secret123"):
        response = self.client.post(
            "/enseignant/connexion",
            json={"identifiant": identifiant, "mot_de_passe": mot_de_passe},
        )
        self.assertEqual(response.status_code, 200)
        return response.json()["token"]

    def _auth(self, token):
        return {"Authorization": f"Bearer {token}"}

    def _creer_classe(self, token, nom="Les Renards", niveau="CE1"):
        response = self.client.post(
            "/classe", json={"nom": nom, "niveau_scolaire": niveau}, headers=self._auth(token)
        )
        self.assertEqual(response.status_code, 201)
        return response.json()

    # ---------- Inscription / connexion enseignant ----------
    def test_inscription_reussie(self) -> None:
        response = self._inscrire()
        self.assertEqual(response.status_code, 201)
        body = response.json()
        self.assertEqual(body["identifiant"], "prof1")
        self.assertIn("ecole_id", body)

    def test_mot_de_passe_jamais_stocke_en_clair(self) -> None:
        self._inscrire(mot_de_passe="motdepasse42")
        # Le hash bcrypt ne contient pas le mot de passe et est verifiable.
        empreinte = comptes.hash_mot_de_passe("motdepasse42")
        self.assertNotIn("motdepasse42", empreinte)
        self.assertTrue(comptes.verifier_mot_de_passe("motdepasse42", empreinte))
        self.assertFalse(comptes.verifier_mot_de_passe("mauvais", empreinte))

    def test_inscription_identifiant_deja_pris(self) -> None:
        self._inscrire()
        response = self._inscrire(nom="Autre")
        self.assertEqual(response.status_code, 409)

    def test_connexion_reussie_donne_un_token(self) -> None:
        self._inscrire()
        response = self.client.post(
            "/enseignant/connexion", json={"identifiant": "prof1", "mot_de_passe": "secret123"}
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["token"])

    def test_connexion_mauvais_mot_de_passe(self) -> None:
        self._inscrire()
        response = self.client.post(
            "/enseignant/connexion", json={"identifiant": "prof1", "mot_de_passe": "faux"}
        )
        self.assertEqual(response.status_code, 401)

    def test_connexion_identifiant_inconnu(self) -> None:
        response = self.client.post(
            "/enseignant/connexion", json={"identifiant": "fantome", "mot_de_passe": "secret123"}
        )
        self.assertEqual(response.status_code, 401)

    # ---------- Classes (protege) ----------
    def test_creer_classe_exige_un_token(self) -> None:
        response = self.client.post("/classe", json={"nom": "X", "niveau_scolaire": "CE1"})
        self.assertEqual(response.status_code, 401)

    def test_creer_et_lister_classes(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, "Les Renards", "CE1")
        self.assertRegex(classe["code_classe"], r"^CE1-[A-Z]+-\d{2}$")

        response = self.client.get("/classe", headers=self._auth(token))
        self.assertEqual(response.status_code, 200)
        classes = response.json()["classes"]
        self.assertEqual(len(classes), 1)
        self.assertEqual(classes[0]["id"], classe["id"])

    def test_creer_classe_niveau_invalide(self) -> None:
        self._inscrire()
        token = self._token()
        response = self.client.post(
            "/classe", json={"nom": "X", "niveau_scolaire": "CE9"}, headers=self._auth(token)
        )
        self.assertEqual(response.status_code, 400)

    # ---------- Eleves ----------
    def test_ajouter_et_retirer_eleve(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        add = self.client.post(
            f"/classe/{classe['id']}/eleve", json={"prenom": "Sofia"}, headers=self._auth(token)
        )
        self.assertEqual(add.status_code, 201)
        eleve_id = add.json()["id"]

        # L'eleve apparait dans la liste publique de la classe.
        rejoindre = self.client.get(f"/classe/rejoindre/{classe['code_classe']}")
        self.assertEqual([e["prenom"] for e in rejoindre.json()["eleves"]], ["Sofia"])

        # Retrait.
        suppr = self.client.delete(
            f"/classe/{classe['id']}/eleve/{eleve_id}", headers=self._auth(token)
        )
        self.assertEqual(suppr.status_code, 204)
        rejoindre2 = self.client.get(f"/classe/rejoindre/{classe['code_classe']}")
        self.assertEqual(rejoindre2.json()["eleves"], [])

    def test_liste_classes_compte_les_eleves(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        for prenom in ("Sofia", "Adam", "Lina"):
            self.client.post(
                f"/classe/{classe['id']}/eleve", json={"prenom": prenom}, headers=self._auth(token)
            )
        classes = self.client.get("/classe", headers=self._auth(token)).json()["classes"]
        self.assertEqual(classes[0]["nb_eleves"], 3)

    def test_lister_eleves_avec_date_ajout(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        self.client.post(
            f"/classe/{classe['id']}/eleve", json={"prenom": "Sofia"}, headers=self._auth(token)
        )
        response = self.client.get(f"/classe/{classe['id']}/eleves", headers=self._auth(token))
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["classe"]["id"], classe["id"])
        self.assertEqual(len(body["eleves"]), 1)
        self.assertEqual(body["eleves"][0]["prenom"], "Sofia")
        self.assertIn("date_creation", body["eleves"][0])

    def test_lister_eleves_exige_le_proprietaire(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe de A", "CE1")

        response = self.client.get(f"/classe/{classeA['id']}/eleves", headers=self._auth(tokenB))
        self.assertEqual(response.status_code, 403)
        # Sans jeton du tout : authentification requise.
        self.assertEqual(self.client.get(f"/classe/{classeA['id']}/eleves").status_code, 401)

    # ---------- Tableau de bord enseignant ----------
    def _seed_progression(self, eleve_id, pattern_name, lecon_id, maitrise, nb_tentatives=1):
        """Insere une ligne de progression directement (sans jouer une session)."""
        with sessionmaker(bind=self.engine, class_=Session)() as db:
            db.add(
                Progression(
                    eleve_id=eleve_id,
                    pattern_name=pattern_name,
                    lecon_id=lecon_id,
                    maitrise=maitrise,
                    nb_tentatives=nb_tentatives,
                )
            )
            db.commit()

    def _ajouter_eleve(self, token, classe_id, prenom):
        return self.client.post(
            f"/classe/{classe_id}/eleve", json={"prenom": prenom}, headers=self._auth(token)
        ).json()["id"]

    def _ajouter_eleve_avec_pin(self, token, classe_id, prenom):
        """Renvoie (eleve_id, pin) : le PIN n'est expose qu'a la creation."""
        corps = self.client.post(
            f"/classe/{classe_id}/eleve", json={"prenom": prenom}, headers=self._auth(token)
        ).json()
        return corps["id"], corps["pin"]

    def test_tableau_de_bord_agrege_par_eleve(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        bon = self._ajouter_eleve(token, classe["id"], "Sofia")
        faible = self._ajouter_eleve(token, classe["id"], "Adam")
        # Sofia maitrise tout ; Adam bloque sur un concept.
        self._seed_progression(bon, "mult_a", "multiplication_division", 3)
        self._seed_progression(bon, "mult_b", "multiplication_division", 3)
        self._seed_progression(faible, "mult_a", "multiplication_division", 3)
        self._seed_progression(faible, "mult_b", "multiplication_division", 1)

        response = self.client.get(
            f"/classe/{classe['id']}/tableau_de_bord", headers=self._auth(token)
        )
        self.assertEqual(response.status_code, 200)
        eleves = {e["prenom"]: e for e in response.json()["eleves"]}
        self.assertEqual(eleves["Sofia"]["nb_acquis"], 2)
        self.assertEqual(eleves["Sofia"]["nb_a_retravailler"], 0)
        self.assertEqual(eleves["Adam"]["nb_acquis"], 1)
        self.assertEqual(eleves["Adam"]["nb_a_retravailler"], 1)
        self.assertEqual(eleves["Adam"]["nb_total"], 2)
        # Le detail des concepts est bien present pour la vue par eleve.
        self.assertEqual(len(eleves["Adam"]["concepts"]), 2)

    def test_tableau_de_bord_inclut_les_eleves_sans_progression(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        self._ajouter_eleve(token, classe["id"], "Nouveau")
        eleves = self.client.get(
            f"/classe/{classe['id']}/tableau_de_bord", headers=self._auth(token)
        ).json()["eleves"]
        self.assertEqual(eleves[0]["nb_total"], 0)
        self.assertEqual(eleves[0]["concepts"], [])

    def test_concepts_difficiles_classe(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        e1 = self._ajouter_eleve(token, classe["id"], "Sofia")
        e2 = self._ajouter_eleve(token, classe["id"], "Adam")
        e3 = self._ajouter_eleve(token, classe["id"], "Lina")
        # "division" bloque 3 eleves (maitrise 1), "posee" en bloque 1.
        for e in (e1, e2, e3):
            self._seed_progression(e, "division_exacte", "multiplication_division", 1)
        self._seed_progression(e1, "mult_posee", "multiplication_division", 1)
        self._seed_progression(e2, "mult_posee", "multiplication_division", 3)  # acquis, non compte

        concepts = self.client.get(
            f"/classe/{classe['id']}/concepts_difficiles", headers=self._auth(token)
        ).json()["concepts"]
        self.assertEqual(concepts[0]["pattern_name"], "division_exacte")
        self.assertEqual(concepts[0]["nb_eleves_en_difficulte"], 3)
        self.assertEqual(concepts[1]["pattern_name"], "mult_posee")
        self.assertEqual(concepts[1]["nb_eleves_en_difficulte"], 1)

    def test_tableau_de_bord_cloisonne_par_enseignant(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe de A", "CE1")

        for chemin in ("tableau_de_bord", "concepts_difficiles"):
            self.assertEqual(
                self.client.get(
                    f"/classe/{classeA['id']}/{chemin}", headers=self._auth(tokenB)
                ).status_code,
                403,
                f"un autre enseignant ne doit pas voir {chemin}",
            )
            self.assertEqual(
                self.client.get(f"/classe/{classeA['id']}/{chemin}").status_code,
                401,
                f"{chemin} exige une authentification",
            )

    # ---------- Cloisonnement entre enseignants ----------
    def test_enseignant_ne_voit_pas_les_classes_d_un_autre(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        self._creer_classe(tokenA, "Classe de A", "CE1")

        listeB = self.client.get("/classe", headers=self._auth(tokenB))
        self.assertEqual(listeB.json()["classes"], [])

    def test_enseignant_ne_peut_pas_ajouter_eleve_a_la_classe_d_un_autre(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe de A", "CE1")

        response = self.client.post(
            f"/classe/{classeA['id']}/eleve",
            json={"prenom": "Intrus"},
            headers=self._auth(tokenB),
        )
        self.assertEqual(response.status_code, 403)

    def test_enseignant_ne_peut_pas_retirer_eleve_d_une_autre_classe(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe de A", "CE1")
        eleve_id = self.client.post(
            f"/classe/{classeA['id']}/eleve", json={"prenom": "Sofia"}, headers=self._auth(tokenA)
        ).json()["id"]

        response = self.client.delete(
            f"/classe/{classeA['id']}/eleve/{eleve_id}", headers=self._auth(tokenB)
        )
        self.assertEqual(response.status_code, 403)

    # ---------- Connexion eleve ----------
    def test_connexion_eleve_reussie(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        eleve_id, pin = self._ajouter_eleve_avec_pin(token, classe["id"], "Sofia")

        response = self.client.post(
            f"/eleve/{eleve_id}/connexion",
            json={"code_classe": classe["code_classe"], "pin": pin},
        )
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["token"])
        self.assertEqual(body["eleve"]["id"], eleve_id)
        self.assertEqual(body["eleve"]["niveau_scolaire"], "CE1")

    def test_creation_eleve_renvoie_un_pin_a_4_chiffres(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        corps = self.client.post(
            f"/classe/{classe['id']}/eleve", json={"prenom": "Sofia"}, headers=self._auth(token)
        ).json()
        self.assertRegex(corps["pin"], r"^\d{4}$")

    def test_pin_affiche_une_seule_fois_a_la_creation(self) -> None:
        # Le PIN n'apparait qu'a la creation : ni la liste enseignant ni la
        # liste publique "rejoindre" ne le reexposent (seul son hash est stocke).
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        eleve_id, _pin = self._ajouter_eleve_avec_pin(token, classe["id"], "Sofia")

        liste = self.client.get(
            f"/classe/{classe['id']}/eleves", headers=self._auth(token)
        ).json()["eleves"]
        self.assertNotIn("pin", liste[0])
        self.assertNotIn("pin_hash", liste[0])

        rejoindre = self.client.get(f"/classe/rejoindre/{classe['code_classe']}").json()
        self.assertNotIn("pin", rejoindre["eleves"][0])
        self.assertNotIn("pin_hash", rejoindre["eleves"][0])
        self.assertEqual(eleve_id, rejoindre["eleves"][0]["id"])

    def test_connexion_eleve_mauvais_pin(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        eleve_id, pin = self._ajouter_eleve_avec_pin(token, classe["id"], "Sofia")
        mauvais = "0000" if pin != "0000" else "1111"

        response = self.client.post(
            f"/eleve/{eleve_id}/connexion",
            json={"code_classe": classe["code_classe"], "pin": mauvais},
        )
        self.assertEqual(response.status_code, 403)

    def test_connexion_eleve_pin_manquant_refuse(self) -> None:
        # Sans PIN, la requete est invalide (422) : le champ est obligatoire.
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        eleve_id, _pin = self._ajouter_eleve_avec_pin(token, classe["id"], "Sofia")
        response = self.client.post(
            f"/eleve/{eleve_id}/connexion", json={"code_classe": classe["code_classe"]}
        )
        self.assertEqual(response.status_code, 422)

    def test_rejoindre_classe_code_inconnu(self) -> None:
        response = self.client.get("/classe/rejoindre/CE1-INEXISTANT-00")
        self.assertEqual(response.status_code, 404)

    def test_connexion_eleve_mauvais_code(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        eleve_id, pin = self._ajouter_eleve_avec_pin(token, classe["id"], "Sofia")

        # PIN correct mais mauvais code -> refuse (l'id doit appartenir au code).
        response = self.client.post(
            f"/eleve/{eleve_id}/connexion", json={"code_classe": "CE1-FAUX-99", "pin": pin}
        )
        self.assertEqual(response.status_code, 403)

    def test_connexion_eleve_id_d_une_autre_classe(self) -> None:
        # Anti-usurpation : un id d'eleve d'une autre classe, avec le code d'ici,
        # doit echouer (l'id doit appartenir a la classe du code fourni).
        self._inscrire()
        token = self._token()
        classe1 = self._creer_classe(token, "Classe 1", "CE1")
        classe2 = self._creer_classe(token, "Classe 2", "CE2")
        eleve_classe1, pin1 = self._ajouter_eleve_avec_pin(token, classe1["id"], "Sofia")

        # id + PIN de l'eleve de la classe 1, mais code de la classe 2 -> refuse.
        response = self.client.post(
            f"/eleve/{eleve_classe1}/connexion",
            json={"code_classe": classe2["code_classe"], "pin": pin1},
        )
        self.assertEqual(response.status_code, 403)

    # ---------- Reinitialisation du PIN eleve ----------
    def test_reinitialiser_pin_invalide_l_ancien_et_active_le_nouveau(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        eleve_id, ancien_pin = self._ajouter_eleve_avec_pin(token, classe["id"], "Sofia")

        # Reinitialisation -> nouveau PIN a 4 chiffres, renvoye une seule fois.
        reinit = self.client.post(
            f"/classe/{classe['id']}/eleve/{eleve_id}/reinitialiser_pin",
            headers=self._auth(token),
        )
        self.assertEqual(reinit.status_code, 200)
        nouveau_pin = reinit.json()["pin"]
        self.assertRegex(nouveau_pin, r"^\d{4}$")
        self.assertNotEqual(nouveau_pin, ancien_pin)

        # L'ancien PIN ne fonctionne plus.
        rejet = self.client.post(
            f"/eleve/{eleve_id}/connexion",
            json={"code_classe": classe["code_classe"], "pin": ancien_pin},
        )
        self.assertEqual(rejet.status_code, 403)

        # Le nouveau PIN fonctionne.
        ok = self.client.post(
            f"/eleve/{eleve_id}/connexion",
            json={"code_classe": classe["code_classe"], "pin": nouveau_pin},
        )
        self.assertEqual(ok.status_code, 200)

    def test_reinitialiser_pin_reserve_au_proprietaire(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe de A", "CE1")
        eleve_id, _pin = self._ajouter_eleve_avec_pin(tokenA, classeA["id"], "Sofia")

        # Un autre enseignant ne peut pas reinitialiser -> 403.
        autre = self.client.post(
            f"/classe/{classeA['id']}/eleve/{eleve_id}/reinitialiser_pin",
            headers=self._auth(tokenB),
        )
        self.assertEqual(autre.status_code, 403)
        # Sans jeton du tout -> authentification requise.
        anon = self.client.post(
            f"/classe/{classeA['id']}/eleve/{eleve_id}/reinitialiser_pin"
        )
        self.assertEqual(anon.status_code, 401)

    def test_reinitialiser_pin_eleve_d_une_autre_classe_introuvable(self) -> None:
        # L'eleve doit appartenir a la classe de l'URL, sinon 404.
        self._inscrire()
        token = self._token()
        classe1 = self._creer_classe(token, "Classe 1", "CE1")
        classe2 = self._creer_classe(token, "Classe 2", "CE2")
        eleve_id, _pin = self._ajouter_eleve_avec_pin(token, classe1["id"], "Sofia")

        response = self.client.post(
            f"/classe/{classe2['id']}/eleve/{eleve_id}/reinitialiser_pin",
            headers=self._auth(token),
        )
        self.assertEqual(response.status_code, 404)

    # ---------- Portail parent (lecture seule) ----------
    def _creer_eleve(self, token, classe_id, prenom):
        """Renvoie le corps de creation (id, pin, code_parent : exposes une seule fois)."""
        return self.client.post(
            f"/classe/{classe_id}/eleve", json={"prenom": prenom}, headers=self._auth(token)
        ).json()

    def test_acces_parent_donne_la_progression_du_bon_eleve(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        eleve = self._creer_eleve(token, classe["id"], "Sofia")
        self._seed_progression(eleve["id"], "division_exacte", "multiplication_division", 2)

        # Le parent echange son code contre un token, puis lit la progression.
        acces = self.client.get(f"/parent/acces/{eleve['code_parent']}")
        self.assertEqual(acces.status_code, 200)
        ptoken = acces.json()["token"]
        self.assertEqual(acces.json()["eleve"]["id"], eleve["id"])
        self.assertEqual(acces.json()["eleve"]["niveau_scolaire"], "CE3")

        prog = self.client.get("/parent/progression", headers=self._auth(ptoken))
        self.assertEqual(prog.status_code, 200)
        corps = prog.json()
        self.assertEqual(corps["eleve"]["id"], eleve["id"])
        self.assertEqual(corps["eleve"]["prenom"], "Sofia")
        self.assertEqual([p["pattern_name"] for p in corps["progression"]], ["division_exacte"])
        self.assertEqual(corps["progression"][0]["maitrise"], 2)

    def test_acces_parent_code_invalide(self) -> None:
        response = self.client.get("/parent/acces/CODEFAUX9")
        self.assertEqual(response.status_code, 403)

    def test_parent_progression_sans_token(self) -> None:
        # /parent/progression sans token parent -> refuse.
        self.assertEqual(self.client.get("/parent/progression").status_code, 401)

    def test_parent_ne_voit_que_son_enfant(self) -> None:
        # Le token du parent de A ne donne QUE la progression de A, jamais celle de B.
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        a = self._creer_eleve(token, classe["id"], "Sofia")
        b = self._creer_eleve(token, classe["id"], "Adam")
        self._seed_progression(a["id"], "concept_a", "multiplication_division", 3)
        self._seed_progression(b["id"], "concept_b", "multiplication_division", 1)

        ptoken_a = self.client.get(f"/parent/acces/{a['code_parent']}").json()["token"]
        corps = self.client.get("/parent/progression", headers=self._auth(ptoken_a)).json()
        self.assertEqual(corps["eleve"]["id"], a["id"])
        patterns = [p["pattern_name"] for p in corps["progression"]]
        self.assertEqual(patterns, ["concept_a"])  # jamais concept_b

    # ---------- Notifications parent (resume + alerte) ----------
    def test_notifications_parent_sans_token(self) -> None:
        # /parent/notifications sans token parent -> refuse (meme protection).
        self.assertEqual(self.client.get("/parent/notifications").status_code, 401)

    def test_notifications_parent_refuse_token_enseignant(self) -> None:
        # Un token enseignant n'est PAS un token parent : acces refuse.
        self._inscrire()
        token = self._token()
        self.assertEqual(
            self.client.get("/parent/notifications", headers=self._auth(token)).status_code, 401
        )

    def test_notifications_parent_reflete_activite_et_alerte(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        eleve = self._creer_eleve(token, classe["id"], "Sofia")
        # Activite recente (date par defaut = maintenant) : un concept bloque
        # (maitrise 1 apres 3 tentatives) doit declencher l'alerte.
        self._seed_progression(eleve["id"], "division_exacte", "mult_div", 1, nb_tentatives=3)

        ptoken = self.client.get(f"/parent/acces/{eleve['code_parent']}").json()["token"]
        reponse = self.client.get("/parent/notifications", headers=self._auth(ptoken))
        self.assertEqual(reponse.status_code, 200)
        corps = reponse.json()
        self.assertEqual(corps["eleve"]["id"], eleve["id"])
        # Le resume reflete l'activite recente.
        travailles = [c["pattern_name"] for c in corps["resume"]["concepts_travailles"]]
        self.assertIn("division_exacte", travailles)
        # L'alerte de blocage est active et cible le bon concept.
        self.assertTrue(corps["alerte"]["active"])
        types = {a["type"] for a in corps["alerte"]["alertes"]}
        self.assertIn("concept_bloque", types)

    def test_notifications_parent_ne_voit_que_son_enfant(self) -> None:
        # Le token du parent de A ne renvoie que les notifications de A.
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        a = self._creer_eleve(token, classe["id"], "Sofia")
        b = self._creer_eleve(token, classe["id"], "Adam")
        self._seed_progression(b["id"], "concept_b", "mult_div", 1, nb_tentatives=5)

        ptoken_a = self.client.get(f"/parent/acces/{a['code_parent']}").json()["token"]
        corps = self.client.get("/parent/notifications", headers=self._auth(ptoken_a)).json()
        self.assertEqual(corps["eleve"]["id"], a["id"])
        self.assertEqual(corps["resume"]["concepts_travailles"], [])  # A n'a rien fait

    # ---------- Rapport IA (endpoints, LLM mocke) ----------
    _RAPPORT_MOCK = "Un texte d'appreciation redige, bienveillant et sans aucun chiffre invente."

    def test_rapport_ia_enseignant_exige_le_proprietaire(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe A", "CE3")
        eleve = self._creer_eleve(tokenA, classeA["id"], "Sofia")

        url = f"/classe/{classeA['id']}/rapport_ia/{eleve['id']}"
        # Sans token : 401 ; token d'un autre enseignant : 403.
        self.assertEqual(self.client.get(url).status_code, 401)
        self.assertEqual(self.client.get(url, headers=self._auth(tokenB)).status_code, 403)

    def test_rapport_ia_enseignant_retourne_le_texte(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        eleve = self._creer_eleve(token, classe["id"], "Sofia")
        self._seed_progression(eleve["id"], "division_exacte", "mult_div", 3)

        with patch("notifications._appel_gemini_rapport", return_value=self._RAPPORT_MOCK):
            reponse = self.client.get(
                f"/classe/{classe['id']}/rapport_ia/{eleve['id']}", headers=self._auth(token)
            )
        self.assertEqual(reponse.status_code, 200)
        corps = reponse.json()
        self.assertEqual(corps["eleve"]["id"], eleve["id"])
        self.assertEqual(corps["rapport"]["source"], "ia")
        self.assertEqual(corps["rapport"]["destinataire"], "enseignant")
        self.assertEqual(corps["rapport"]["texte"], self._RAPPORT_MOCK)

    def test_rapport_ia_parent_sans_token(self) -> None:
        self.assertEqual(self.client.get("/parent/rapport_ia").status_code, 401)

    def test_rapport_ia_parent_retourne_le_texte(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        eleve = self._creer_eleve(token, classe["id"], "Sofia")
        self._seed_progression(eleve["id"], "division_exacte", "mult_div", 2)
        ptoken = self.client.get(f"/parent/acces/{eleve['code_parent']}").json()["token"]

        with patch("notifications._appel_gemini_rapport", return_value=self._RAPPORT_MOCK):
            reponse = self.client.get("/parent/rapport_ia", headers=self._auth(ptoken))
        self.assertEqual(reponse.status_code, 200)
        corps = reponse.json()
        self.assertEqual(corps["eleve"]["id"], eleve["id"])
        self.assertEqual(corps["rapport"]["destinataire"], "parent")
        self.assertEqual(corps["rapport"]["texte"], self._RAPPORT_MOCK)

    # ---------- Durcissement du niveau cote serveur ----------
    def test_session_eleve_forcee_au_niveau_de_sa_classe(self) -> None:
        # Un eleve d'une classe CE1 qui demande une session en CE5 doit jouer en
        # CE1 : le niveau envoye par le client est ignore au profit de la base.
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE1")
        eleve = self._creer_eleve(token, classe["id"], "Sofia")
        tel = self.client.post(
            f"/eleve/{eleve['id']}/connexion",
            json={"code_classe": classe["code_classe"], "pin": eleve["pin"]},
        ).json()["token"]

        reponse = self.client.post(
            "/session/demarrer", json={"niveau_scolaire": "CE5"}, headers=self._auth(tel)
        )
        self.assertEqual(reponse.status_code, 200)
        # Le niveau reellement utilise est celui de la classe (CE1), pas CE5.
        self.assertEqual(reponse.json()["progression"]["niveau_scolaire"], "CE1")

        # Controle : en mode invite (sans token), le niveau demande est respecte.
        invite = self.client.post("/session/demarrer", json={"niveau_scolaire": "CE5"})
        self.assertEqual(invite.status_code, 200)
        self.assertEqual(invite.json()["progression"]["niveau_scolaire"], "CE5")

    def test_regenerer_code_parent_invalide_l_ancien(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        eleve = self._creer_eleve(token, classe["id"], "Sofia")
        ancien = eleve["code_parent"]

        regen = self.client.post(
            f"/classe/{classe['id']}/eleve/{eleve['id']}/code_parent", headers=self._auth(token)
        )
        self.assertEqual(regen.status_code, 200)
        nouveau = regen.json()["code_parent"]
        self.assertNotEqual(nouveau, ancien)
        self.assertRegex(nouveau, r"^[A-Z0-9]{8}$")

        # L'ancien code ne donne plus acces, le nouveau si.
        self.assertEqual(self.client.get(f"/parent/acces/{ancien}").status_code, 403)
        self.assertEqual(self.client.get(f"/parent/acces/{nouveau}").status_code, 200)

    def test_regenerer_code_parent_reserve_au_proprietaire(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe de A", "CE1")
        eleve = self._creer_eleve(tokenA, classeA["id"], "Sofia")

        autre = self.client.post(
            f"/classe/{classeA['id']}/eleve/{eleve['id']}/code_parent", headers=self._auth(tokenB)
        )
        self.assertEqual(autre.status_code, 403)
        anon = self.client.post(f"/classe/{classeA['id']}/eleve/{eleve['id']}/code_parent")
        self.assertEqual(anon.status_code, 401)

    def test_token_parent_ne_permet_aucune_ecriture(self) -> None:
        # Le token parent est en lecture seule : il ne doit ouvrir aucun endpoint
        # d'ecriture / de gestion, ni la progression brute par id d'eleve.
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        eleve = self._creer_eleve(token, classe["id"], "Sofia")
        ptoken = self.client.get(f"/parent/acces/{eleve['code_parent']}").json()["token"]
        ph = self._auth(ptoken)

        # Gestion de classe (enseignant) -> 401 (token non enseignant).
        self.assertEqual(
            self.client.post(
                f"/classe/{classe['id']}/eleve", json={"prenom": "Intrus"}, headers=ph
            ).status_code,
            401,
        )
        self.assertEqual(
            self.client.post(
                f"/classe/{classe['id']}/eleve/{eleve['id']}/reinitialiser_pin", headers=ph
            ).status_code,
            401,
        )
        self.assertEqual(
            self.client.post(
                f"/classe/{classe['id']}/eleve/{eleve['id']}/code_parent", headers=ph
            ).status_code,
            401,
        )
        self.assertEqual(
            self.client.delete(
                f"/classe/{classe['id']}/eleve/{eleve['id']}", headers=ph
            ).status_code,
            401,
        )
        self.assertEqual(self.client.get("/classe", headers=ph).status_code, 401)
        # Progression brute par id d'eleve (token eleve/enseignant only) -> 403.
        self.assertEqual(
            self.client.get(f"/eleve/{eleve['id']}/progression", headers=ph).status_code, 403
        )

    # ---------- Export Excel de la classe ----------
    def _ouvrir_classeur(self, contenu: bytes):
        import io

        from openpyxl import load_workbook

        return load_workbook(io.BytesIO(contenu))

    def test_export_excel_est_un_classeur_valide(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        sofia = self._ajouter_eleve(token, classe["id"], "Sofia")
        adam = self._ajouter_eleve(token, classe["id"], "Adam")
        # Concepts differents entre les deux eleves (3 concepts distincts au total).
        self._seed_progression(sofia, "mult_a", "multiplication_division", 3)
        self._seed_progression(sofia, "mult_b", "multiplication_division", 2)
        self._seed_progression(adam, "mult_a", "multiplication_division", 1)
        self._seed_progression(adam, "div_a", "multiplication_division", 3)

        reponse = self.client.get(
            f"/classe/{classe['id']}/export_excel", headers=self._auth(token)
        )
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(
            reponse.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", reponse.headers["content-disposition"])
        self.assertIn(".xlsx", reponse.headers["content-disposition"])

        # Le fichier s'ouvre reellement comme un classeur Excel.
        wb = self._ouvrir_classeur(reponse.content)
        self.assertEqual(wb.sheetnames, ["Vue d'ensemble", "Detail"])

        # --- Feuille "Vue d'ensemble" : croisement eleve x concept ---
        vue = wb["Vue d'ensemble"]
        self.assertTrue(str(vue["A1"].value).startswith("Classe :"))
        self.assertIn("CE3", str(vue["A2"].value))
        self.assertIn(classe["code_classe"], str(vue["A2"].value))
        self.assertEqual(vue.cell(row=5, column=1).value, "Eleve")
        # 3 concepts distincts -> 3 colonnes (B, C, D), la 4e (E) vide.
        entetes = [vue.cell(row=5, column=c).value for c in range(2, 6)]
        self.assertEqual([e for e in entetes if e], ["div a", "mult a", "mult b"])
        # Eleves tries par prenom : Adam (ligne 6), Sofia (ligne 7).
        self.assertEqual(vue.cell(row=6, column=1).value, "Adam")
        self.assertEqual(vue.cell(row=7, column=1).value, "Sofia")
        # Adam : div_a=3 (col B), mult_a=1 (col C), mult_b non traverse (vide).
        self.assertEqual(vue.cell(row=6, column=2).value, 3)
        self.assertEqual(vue.cell(row=6, column=3).value, 1)
        self.assertIsNone(vue.cell(row=6, column=4).value)
        # Couleur de fond conditionnelle : maitrise 1 -> rouge clair (FFC7CE).
        self.assertTrue(str(vue.cell(row=6, column=3).fill.fgColor.rgb).endswith("FFC7CE"))
        # Maitrise 3 -> vert (C6EFCE).
        self.assertTrue(str(vue.cell(row=6, column=2).fill.fgColor.rgb).endswith("C6EFCE"))

        # --- Feuille "Detail" : une ligne par eleve x concept ---
        detail = wb["Detail"]
        self.assertEqual(
            [detail.cell(row=5, column=c).value for c in range(1, 7)],
            ["Eleve", "Lecon", "Concept", "Maitrise", "Niveau", "Date"],
        )
        # 4 lignes de donnees (2 concepts par eleve).
        lignes_detail = [
            detail.cell(row=r, column=1).value
            for r in range(6, detail.max_row + 1)
            if detail.cell(row=r, column=1).value
        ]
        self.assertEqual(len(lignes_detail), 4)

    def test_export_excel_classe_vide(self) -> None:
        # Une classe sans eleve produit tout de meme un classeur valide.
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        reponse = self.client.get(
            f"/classe/{classe['id']}/export_excel", headers=self._auth(token)
        )
        self.assertEqual(reponse.status_code, 200)
        wb = self._ouvrir_classeur(reponse.content)
        self.assertEqual(wb.sheetnames, ["Vue d'ensemble", "Detail"])
        # En-tete present, aucune ligne d'eleve.
        self.assertEqual(wb["Vue d'ensemble"].cell(row=5, column=1).value, "Eleve")
        self.assertIsNone(wb["Vue d'ensemble"].cell(row=6, column=1).value)

    def test_export_excel_reserve_au_proprietaire(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe de A", "CE1")

        self.assertEqual(
            self.client.get(
                f"/classe/{classeA['id']}/export_excel", headers=self._auth(tokenB)
            ).status_code,
            403,
        )
        self.assertEqual(
            self.client.get(f"/classe/{classeA['id']}/export_excel").status_code, 401
        )

    # ---------- Assignations ----------
    def test_assigner_lecon_cree_une_ligne_par_eleve(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        e1 = self._ajouter_eleve(token, classe["id"], "Sofia")
        e2 = self._ajouter_eleve(token, classe["id"], "Adam")

        rep = self.client.post(
            f"/classe/{classe['id']}/assigner",
            json={"eleve_ids": [e1, e2], "lecon_id": "multiplication_division"},
            headers=self._auth(token),
        )
        self.assertEqual(rep.status_code, 201)
        assignations = rep.json()["assignations"]
        self.assertEqual(len(assignations), 2)
        self.assertTrue(all(a["type"] == "lecon" for a in assignations))
        self.assertTrue(all(a["lecon_id"] == "multiplication_division" for a in assignations))
        self.assertTrue(all(a["terminee"] is False for a in assignations))

    def test_assigner_revision_ciblee(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        e1 = self._ajouter_eleve(token, classe["id"], "Sofia")
        rep = self.client.post(
            f"/classe/{classe['id']}/assigner",
            json={"eleve_ids": [e1], "patterns": ["division_exacte_partage"]},
            headers=self._auth(token),
        )
        self.assertEqual(rep.status_code, 201)
        a = rep.json()["assignations"][0]
        self.assertEqual(a["type"], "revision")
        self.assertEqual(a["patterns"], ["division_exacte_partage"])
        self.assertIsNone(a["lecon_id"])

    def test_assigner_exige_lecon_ou_patterns_exclusif(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token)
        e1 = self._ajouter_eleve(token, classe["id"], "Sofia")
        # Ni l'un ni l'autre -> 400.
        self.assertEqual(
            self.client.post(
                f"/classe/{classe['id']}/assigner",
                json={"eleve_ids": [e1]},
                headers=self._auth(token),
            ).status_code,
            400,
        )
        # Les deux a la fois -> 400.
        self.assertEqual(
            self.client.post(
                f"/classe/{classe['id']}/assigner",
                json={"eleve_ids": [e1], "lecon_id": "x", "patterns": ["y"]},
                headers=self._auth(token),
            ).status_code,
            400,
        )

    def test_assigner_cloisonne_par_enseignant(self) -> None:
        self._inscrire("profA", "secretA")
        self._inscrire("profB", "secretB")
        tokenA = self._token("profA", "secretA")
        tokenB = self._token("profB", "secretB")
        classeA = self._creer_classe(tokenA, "Classe de A", "CE1")
        classeB = self._creer_classe(tokenB, "Classe de B", "CE1")
        eleveA = self._ajouter_eleve(tokenA, classeA["id"], "Sofia")
        eleveB = self._ajouter_eleve(tokenB, classeB["id"], "Adam")

        # B ne peut pas assigner dans la classe de A -> 403.
        self.assertEqual(
            self.client.post(
                f"/classe/{classeA['id']}/assigner",
                json={"eleve_ids": [eleveA], "lecon_id": "l"},
                headers=self._auth(tokenB),
            ).status_code,
            403,
        )
        # A ne peut pas assigner a un eleve qui n'est pas dans SA classe -> 400.
        self.assertEqual(
            self.client.post(
                f"/classe/{classeA['id']}/assigner",
                json={"eleve_ids": [eleveA, eleveB], "lecon_id": "l"},
                headers=self._auth(tokenA),
            ).status_code,
            400,
        )

    def test_eleve_ne_voit_que_ses_assignations(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        a = self._creer_eleve(token, classe["id"], "Sofia")
        b = self._creer_eleve(token, classe["id"], "Adam")
        self.client.post(
            f"/classe/{classe['id']}/assigner",
            json={"eleve_ids": [a["id"]], "lecon_id": "multiplication_division"},
            headers=self._auth(token),
        )
        # Token eleve de A : voit sa (une) assignation, pas celle des autres.
        tok_a = self.client.post(
            f"/eleve/{a['id']}/connexion",
            json={"code_classe": classe["code_classe"], "pin": a["pin"]},
        ).json()["token"]
        mine = self.client.get(f"/eleve/{a['id']}/assignations", headers=self._auth(tok_a))
        self.assertEqual(mine.status_code, 200)
        self.assertEqual(len(mine.json()["assignations"]), 1)
        # A ne peut pas lire les assignations de B (autre eleve) -> 403.
        self.assertEqual(
            self.client.get(f"/eleve/{b['id']}/assignations", headers=self._auth(tok_a)).status_code,
            403,
        )
        # L'enseignant de la classe peut, lui, lire celles de B.
        self.assertEqual(
            self.client.get(f"/eleve/{b['id']}/assignations", headers=self._auth(token)).status_code,
            200,
        )

    def test_statut_assignations_classe(self) -> None:
        self._inscrire()
        token = self._token()
        classe = self._creer_classe(token, niveau="CE3")
        e1 = self._ajouter_eleve(token, classe["id"], "Sofia")
        self.client.post(
            f"/classe/{classe['id']}/assigner",
            json={"eleve_ids": [e1], "lecon_id": "multiplication_division"},
            headers=self._auth(token),
        )
        rep = self.client.get(f"/classe/{classe['id']}/assignations", headers=self._auth(token))
        self.assertEqual(rep.status_code, 200)
        lignes = rep.json()["assignations"]
        self.assertEqual(len(lignes), 1)
        self.assertEqual(lignes[0]["prenom"], "Sofia")
        self.assertFalse(lignes[0]["terminee"])
        # Cloisonnement : un autre enseignant ne voit pas le statut, et sans jeton 401.
        self._inscrire("autre", "secret999")
        tok2 = self._token("autre", "secret999")
        self.assertEqual(
            self.client.get(
                f"/classe/{classe['id']}/assignations", headers=self._auth(tok2)
            ).status_code,
            403,
        )
        self.assertEqual(
            self.client.get(f"/classe/{classe['id']}/assignations").status_code, 401
        )


class IsolationDeuxEcolesTests(unittest.TestCase):
    """Isolation MULTI-ECOLES : deux ecoles reellement distinctes, un enseignant,
    une classe et un eleve dans chacune. On verifie qu'AUCUNE donnee d'une ecole
    ne fuit vers l'autre, sur TOUS les endpoints proteges (pas un echantillon).

    Distinct du cloisonnement enseignant<->enseignant deja teste : ici les deux
    enseignants relevent d'ecoles differentes, ce que l'ancien modele (ecole
    implicite unique partagee) rendait impossible a exprimer."""

    def setUp(self) -> None:
        self.engine = create_db_engine("sqlite://")
        init_db(self.engine)
        self.testing_session = sessionmaker(
            bind=self.engine, autoflush=False, expire_on_commit=False, class_=Session
        )

        def override_get_db():
            db = self.testing_session()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[comptes.get_db] = override_get_db
        comptes._reset_tokens()
        self.client = TestClient(app)

        # --- Ecole A ---
        self.ecole_a_id = self.client.post(
            "/enseignant/inscription",
            json={"nom": "Prof A", "identifiant": "profA", "mot_de_passe": "secretA", "ecole": "Ecole A"},
        ).json()["ecole_id"]
        self.tokenA = self.client.post(
            "/enseignant/connexion", json={"identifiant": "profA", "mot_de_passe": "secretA"}
        ).json()["token"]
        self.classeA = self.client.post(
            "/classe", json={"nom": "Classe A", "niveau_scolaire": "CE3"}, headers=self._auth(self.tokenA)
        ).json()
        self.eleveA = self.client.post(
            f"/classe/{self.classeA['id']}/eleve", json={"prenom": "Anna"}, headers=self._auth(self.tokenA)
        ).json()
        self._seed_progression(self.eleveA["id"], "division_exacte", "multiplication_division", 1)

        # --- Ecole B ---
        self.ecole_b_id = self.client.post(
            "/enseignant/inscription",
            json={"nom": "Prof B", "identifiant": "profB", "mot_de_passe": "secretB", "ecole": "Ecole B"},
        ).json()["ecole_id"]
        self.tokenB = self.client.post(
            "/enseignant/connexion", json={"identifiant": "profB", "mot_de_passe": "secretB"}
        ).json()["token"]
        self.classeB = self.client.post(
            "/classe", json={"nom": "Classe B", "niveau_scolaire": "CE3"}, headers=self._auth(self.tokenB)
        ).json()
        self.eleveB = self.client.post(
            f"/classe/{self.classeB['id']}/eleve", json={"prenom": "Bob"}, headers=self._auth(self.tokenB)
        ).json()

    def tearDown(self) -> None:
        app.dependency_overrides.clear()
        comptes._reset_tokens()
        self.engine.dispose()

    def _auth(self, token):
        return {"Authorization": f"Bearer {token}"}

    def _seed_progression(self, eleve_id, pattern_name, lecon_id, maitrise, nb_tentatives=1):
        with self.testing_session() as db:
            db.add(
                Progression(
                    eleve_id=eleve_id,
                    pattern_name=pattern_name,
                    lecon_id=lecon_id,
                    maitrise=maitrise,
                    nb_tentatives=nb_tentatives,
                )
            )
            db.commit()

    # --- Le coeur : deux ecoles reellement distinctes ---
    def test_deux_inscriptions_creent_deux_ecoles_distinctes(self) -> None:
        # Le bug historique : toutes les inscriptions retombaient sur une seule
        # ecole implicite. Desormais chaque enseignant fonde la sienne.
        self.assertNotEqual(self.ecole_a_id, self.ecole_b_id)
        with self.testing_session() as db:
            from database import Ecole
            self.assertEqual(db.query(Ecole).count(), 2)

    # --- GET /classe : chaque enseignant ne voit que SON ecole ---
    def test_liste_classes_cloisonnee_par_ecole(self) -> None:
        noms_a = [c["nom"] for c in self.client.get("/classe", headers=self._auth(self.tokenA)).json()["classes"]]
        noms_b = [c["nom"] for c in self.client.get("/classe", headers=self._auth(self.tokenB)).json()["classes"]]
        self.assertEqual(noms_a, ["Classe A"])
        self.assertEqual(noms_b, ["Classe B"])

    # --- Tous les endpoints classe-scopes : B ne touche pas la classe de A ---
    def test_endpoints_classe_scopes_refusent_l_autre_ecole(self) -> None:
        cid = self.classeA["id"]
        eid = self.eleveA["id"]
        b = self._auth(self.tokenB)
        # (methode, url, statut attendu quand B vise l'ecole A)
        cas_403 = [
            ("get", f"/classe/{cid}/eleves"),
            ("get", f"/classe/{cid}/tableau_de_bord"),
            ("get", f"/classe/{cid}/concepts_difficiles"),
            ("get", f"/classe/{cid}/export_excel"),
            ("get", f"/classe/{cid}/assignations"),
            ("get", f"/classe/{cid}/rapport_ia/{eid}"),
            ("post", f"/classe/{cid}/eleve/{eid}/reinitialiser_pin"),
            ("post", f"/classe/{cid}/eleve/{eid}/code_parent"),
            ("delete", f"/classe/{cid}/eleve/{eid}"),
        ]
        for methode, url in cas_403:
            reponse = getattr(self.client, methode)(url, headers=b)
            self.assertEqual(reponse.status_code, 403, f"{methode.upper()} {url} devrait etre 403 pour l'ecole B")

        # POST creation d'eleve dans la classe de A
        self.assertEqual(
            self.client.post(f"/classe/{cid}/eleve", json={"prenom": "Intrus"}, headers=b).status_code, 403
        )
        # POST assignation dans la classe de A
        self.assertEqual(
            self.client.post(
                f"/classe/{cid}/assigner", json={"eleve_ids": [eid], "lecon_id": "multiplication_division"}, headers=b
            ).status_code,
            403,
        )

    # --- Endpoints eleve-par-id : l'enseignant de B ne lit pas un eleve de A ---
    def test_eleve_par_id_refuse_enseignant_autre_ecole(self) -> None:
        b = self._auth(self.tokenB)
        for url in (
            f"/eleve/{self.eleveA['id']}/progression",
            f"/eleve/{self.eleveA['id']}/assignations",
        ):
            self.assertEqual(self.client.get(url, headers=b).status_code, 403, url)

    # --- Assignation croisee interdite meme depuis SA propre classe ---
    def test_assigner_eleve_d_une_autre_ecole_refuse(self) -> None:
        # A vise, depuis SA classe, un eleve de l'ecole B -> 400 (eleve hors classe).
        reponse = self.client.post(
            f"/classe/{self.classeA['id']}/assigner",
            json={"eleve_ids": [self.eleveB["id"]], "lecon_id": "multiplication_division"},
            headers=self._auth(self.tokenA),
        )
        self.assertEqual(reponse.status_code, 400)

    # --- Le code parent d'un eleve de B ne donne acces qu'a B ---
    def test_code_parent_ne_franchit_pas_l_ecole(self) -> None:
        # Le parent de B accede a B, jamais aux donnees de A.
        ptoken_b = self.client.get(f"/parent/acces/{self.eleveB['code_parent']}").json()["token"]
        corps = self.client.get("/parent/progression", headers=self._auth(ptoken_b)).json()
        self.assertEqual(corps["eleve"]["id"], self.eleveB["id"])
        self.assertNotEqual(corps["eleve"]["id"], self.eleveA["id"])

    # --- L'eleve de B ne lit pas la progression/assignations d'un eleve de A ---
    def test_eleve_connecte_ne_lit_pas_l_autre_ecole(self) -> None:
        tok_b = self.client.post(
            f"/eleve/{self.eleveB['id']}/connexion",
            json={"code_classe": self.classeB["code_classe"], "pin": self.eleveB["pin"]},
        ).json()["token"]
        for url in (
            f"/eleve/{self.eleveA['id']}/progression",
            f"/eleve/{self.eleveA['id']}/assignations",
            f"/eleve/{self.eleveA['id']}/personnage",
        ):
            self.assertEqual(self.client.get(url, headers=self._auth(tok_b)).status_code, 403, url)


class RoleAdministrateurTests(unittest.TestCase):
    """Role administrateur d'ecole : vue etablissement + gestion des comptes.

    Scenario de base (setUp) : une ecole A avec un administrateur et DEUX
    enseignants simples (rejoints par invitation), chacun avec sa classe."""

    def setUp(self) -> None:
        self.engine = create_db_engine("sqlite://")
        init_db(self.engine)
        testing_session = sessionmaker(
            bind=self.engine, autoflush=False, expire_on_commit=False, class_=Session
        )

        def override_get_db():
            db = testing_session()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[comptes.get_db] = override_get_db
        comptes._reset_tokens()
        self.client = TestClient(app)

        # Admin de l'ecole A (1er compte sans code -> administrateur).
        self.admin = self._inscrire("admin1", "secretA", "Admin A")
        self.ecole_a = self.admin["ecole_id"]
        self.tAdmin = self._token("admin1", "secretA")

        # Deux enseignants simples rejoignent l'ecole A par invitation.
        code1 = self._inviter(self.tAdmin)
        code2 = self._inviter(self.tAdmin)
        self.prof1 = self._inscrire("prof1", "secret1", "Prof Un", code_invitation=code1)
        self.prof2 = self._inscrire("prof2", "secret2", "Prof Deux", code_invitation=code2)
        self.t1 = self._token("prof1", "secret1")
        self.t2 = self._token("prof2", "secret2")

        # Une classe chacun (l'admin peut aussi avoir la sienne).
        self.classe1 = self._creer_classe(self.t1, "Classe de Un")
        self.classe2 = self._creer_classe(self.t2, "Classe de Deux")

    def tearDown(self) -> None:
        app.dependency_overrides.clear()
        comptes._reset_tokens()
        self.engine.dispose()

    # --- Helpers ---
    def _auth(self, token):
        return {"Authorization": f"Bearer {token}"}

    def _inscrire(self, identifiant, mot_de_passe, nom, code_invitation=None):
        corps = {"nom": nom, "identifiant": identifiant, "mot_de_passe": mot_de_passe}
        if code_invitation is not None:
            corps["code_invitation"] = code_invitation
        response = self.client.post("/enseignant/inscription", json=corps)
        self.assertEqual(response.status_code, 201, response.text)
        return response.json()

    def _token(self, identifiant, mot_de_passe):
        return self.client.post(
            "/enseignant/connexion",
            json={"identifiant": identifiant, "mot_de_passe": mot_de_passe},
        ).json()["token"]

    def _inviter(self, token, email=None):
        response = self.client.post(
            "/ecole/enseignants/inviter", json={"email": email}, headers=self._auth(token)
        )
        self.assertEqual(response.status_code, 201, response.text)
        return response.json()["code"]

    def _creer_classe(self, token, nom, niveau="CE2"):
        return self.client.post(
            "/classe", json={"nom": nom, "niveau_scolaire": niveau}, headers=self._auth(token)
        ).json()

    # --- Modele de role ---
    def test_premier_compte_est_administrateur(self) -> None:
        self.assertEqual(self.admin["role"], "administrateur")
        # Connexion : le role est aussi renvoye (le frontend en depend).
        corps = self.client.post(
            "/enseignant/connexion", json={"identifiant": "admin1", "mot_de_passe": "secretA"}
        ).json()
        self.assertEqual(corps["enseignant"]["role"], "administrateur")

    def test_invite_est_enseignant_simple_dans_la_meme_ecole(self) -> None:
        self.assertEqual(self.prof1["role"], "enseignant")
        self.assertEqual(self.prof1["ecole_id"], self.ecole_a)  # meme ecole que l'admin
        self.assertEqual(self.prof2["role"], "enseignant")

    def test_code_invitation_est_a_usage_unique(self) -> None:
        code = self._inviter(self.tAdmin)
        self._inscrire("prof3", "secret3", "Prof Trois", code_invitation=code)
        # Reutiliser le meme code -> refuse.
        rejeu = self.client.post(
            "/enseignant/inscription",
            json={"nom": "Prof Quatre", "identifiant": "prof4", "mot_de_passe": "secret4",
                  "code_invitation": code},
        )
        self.assertEqual(rejeu.status_code, 400)

    def test_code_invitation_inconnu_refuse(self) -> None:
        rep = self.client.post(
            "/enseignant/inscription",
            json={"nom": "X", "identifiant": "profx", "mot_de_passe": "secretx",
                  "code_invitation": "ECOLE-INCONNU9"},
        )
        self.assertEqual(rep.status_code, 400)

    # --- Vue etablissement (admin voit tout) ---
    def test_admin_voit_toutes_les_classes_de_l_ecole(self) -> None:
        rep = self.client.get("/ecole/classes", headers=self._auth(self.tAdmin))
        self.assertEqual(rep.status_code, 200)
        classes = rep.json()["classes"]
        noms = {c["nom"] for c in classes}
        self.assertEqual(noms, {"Classe de Un", "Classe de Deux"})
        # L'enseignant responsable est indique pour chaque classe.
        par_nom = {c["nom"]: c["enseignant"]["nom"] for c in classes}
        self.assertEqual(par_nom["Classe de Un"], "Prof Un")
        self.assertEqual(par_nom["Classe de Deux"], "Prof Deux")

    def test_enseignant_simple_ne_voit_que_sa_classe(self) -> None:
        # La vue enseignant normale (GET /classe) reste cloisonnee.
        liste1 = self.client.get("/classe", headers=self._auth(self.t1)).json()["classes"]
        self.assertEqual([c["nom"] for c in liste1], ["Classe de Un"])
        liste2 = self.client.get("/classe", headers=self._auth(self.t2)).json()["classes"]
        self.assertEqual([c["nom"] for c in liste2], ["Classe de Deux"])

    def test_admin_liste_les_enseignants(self) -> None:
        rep = self.client.get("/ecole/enseignants", headers=self._auth(self.tAdmin))
        self.assertEqual(rep.status_code, 200)
        gens = {e["identifiant"]: e for e in rep.json()["enseignants"]}
        self.assertEqual(set(gens), {"admin1", "prof1", "prof2"})
        self.assertEqual(gens["admin1"]["role"], "administrateur")
        self.assertEqual(gens["prof1"]["role"], "enseignant")
        self.assertTrue(gens["admin1"]["est_moi"])
        self.assertEqual(gens["prof2"]["nb_classes"], 1)

    # --- Endpoints admin fermes aux enseignants simples ---
    def test_endpoints_admin_interdits_a_l_enseignant_simple(self) -> None:
        chemins = [
            ("get", "/ecole/classes", None),
            ("get", "/ecole/enseignants", None),
            ("post", "/ecole/enseignants/inviter", {"email": None}),
            ("put", f"/ecole/enseignants/{self.prof2['id']}/role", {"role": "administrateur"}),
        ]
        for methode, url, corps in chemins:
            appel = getattr(self.client, methode)
            kwargs = {"headers": self._auth(self.t1)}
            if corps is not None:
                kwargs["json"] = corps
            self.assertEqual(appel(url, **kwargs).status_code, 403, f"{methode} {url}")
            # Sans jeton : 401.
            kwargs.pop("headers")
            self.assertEqual(appel(url, **kwargs).status_code, 401, f"{methode} {url} anon")

    # --- Promotion / retrogradation + garde-fou dernier admin ---
    def test_promotion_et_retrogradation(self) -> None:
        # Promeut prof1 administrateur.
        rep = self.client.put(
            f"/ecole/enseignants/{self.prof1['id']}/role",
            json={"role": "administrateur"}, headers=self._auth(self.tAdmin),
        )
        self.assertEqual(rep.status_code, 200)
        self.assertEqual(rep.json()["role"], "administrateur")
        # prof1 (desormais admin) peut voir la vue etablissement.
        self.assertEqual(
            self.client.get("/ecole/classes", headers=self._auth(self.t1)).status_code, 200
        )
        # On peut maintenant retrograder l'admin d'origine (il reste prof1 admin).
        rep2 = self.client.put(
            f"/ecole/enseignants/{self.admin['id']}/role",
            json={"role": "enseignant"}, headers=self._auth(self.tAdmin),
        )
        self.assertEqual(rep2.status_code, 200)
        self.assertEqual(rep2.json()["role"], "enseignant")

    def test_dernier_admin_ne_peut_pas_se_retrograder(self) -> None:
        # admin1 est le seul administrateur : se retrograder est refuse (409).
        rep = self.client.put(
            f"/ecole/enseignants/{self.admin['id']}/role",
            json={"role": "enseignant"}, headers=self._auth(self.tAdmin),
        )
        self.assertEqual(rep.status_code, 409)
        # Il reste bien administrateur.
        self.assertEqual(
            self.client.get("/ecole/classes", headers=self._auth(self.tAdmin)).status_code, 200
        )

    # --- Cloisonnement inter-ecoles du role admin ---
    def test_admin_ne_gere_pas_une_autre_ecole(self) -> None:
        # Ecole B, son propre admin + un enseignant.
        self._inscrire("adminB", "secretB", "Admin B")
        tB = self._token("adminB", "secretB")
        codeB = self._inviter(tB)
        profB = self._inscrire("profB", "secretpb", "Prof B", code_invitation=codeB)
        self._creer_classe(tB, "Classe B")

        # L'admin de A ne voit QUE les classes de A.
        noms_a = {c["nom"] for c in self.client.get(
            "/ecole/classes", headers=self._auth(self.tAdmin)).json()["classes"]}
        self.assertNotIn("Classe B", noms_a)
        # L'admin de A ne peut pas changer le role d'un enseignant de B -> 404.
        rep = self.client.put(
            f"/ecole/enseignants/{profB['id']}/role",
            json={"role": "administrateur"}, headers=self._auth(self.tAdmin),
        )
        self.assertEqual(rep.status_code, 404)


if __name__ == "__main__":
    unittest.main()
