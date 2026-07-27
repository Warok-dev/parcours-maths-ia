"""Generation a la volee du classeur Excel de progression d'une classe.

Isole de l'endpoint (comptes.py) : ce module ne fait que construire un
openpyxl.Workbook a partir de la meme structure de donnees que le tableau de
bord enseignant, ce qui le rend testable sans HTTP ni base. Aucun fichier n'est
ecrit sur disque : l'appelant serialise le classeur en memoire.

Structure d'entree attendue (identique au tableau de bord) :
    classe = {"nom": str, "niveau_scolaire": str, "code_classe": str}
    eleves = [
        {
            "prenom": str,
            "concepts": [
                {"pattern_name": str, "lecon_id": str | None,
                 "maitrise": int, "date_derniere_tentative": str},
                ...
            ],
        },
        ...
    ]
"""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Couleurs de fond par niveau de maitrise (palette classique d'Excel) :
# 1 = a retravailler (rouge clair), 2 = en bonne voie (jaune), 3 = acquis (vert).
_FILLS_MAITRISE = {
    1: PatternFill("solid", fgColor="FFC7CE"),
    2: PatternFill("solid", fgColor="FFEB9C"),
    3: PatternFill("solid", fgColor="C6EFCE"),
}
_LIBELLES_MAITRISE = {1: "A retravailler", 2: "En bonne voie", 3: "Acquis"}

_TITRE_FONT = Font(bold=True, size=14)
_ENTETE_FONT = Font(bold=True)
_CENTRE = Alignment(horizontal="center", vertical="center")


def _libelle_concept(pattern_name: str) -> str:
    """Rend un pattern_name un peu plus lisible (underscores -> espaces).
    On garde le pattern brut comme source pour ne pas creer d'ambiguite."""
    return str(pattern_name or "").replace("_", " ")


def _ecrire_entete_feuille(ws: Worksheet, classe: dict, date_export: str) -> int:
    """Ecrit le bandeau d'en-tete (classe, niveau, date) et renvoie l'index de
    la premiere ligne libre pour le tableau."""
    ws["A1"] = f"Classe : {classe.get('nom', '')}"
    ws["A1"].font = _TITRE_FONT
    ws["A2"] = f"Niveau : {classe.get('niveau_scolaire', '')}    Code : {classe.get('code_classe', '')}"
    ws["A3"] = f"Export du {date_export}"
    ws["A3"].font = Font(italic=True, color="808080")
    return 5  # ligne 4 laissee vide, le tableau commence en ligne 5


def _concepts_de_la_classe(eleves: list[dict]) -> list[dict]:
    """Liste ordonnee et dedupliquee des concepts traverses par la classe.
    Ordre stable : par lecon puis par nom de concept."""
    vus: dict[str, dict] = {}
    for eleve in eleves:
        for c in eleve.get("concepts", []):
            pattern = c.get("pattern_name")
            if pattern and pattern not in vus:
                vus[pattern] = {"pattern_name": pattern, "lecon_id": c.get("lecon_id") or ""}
    return sorted(vus.values(), key=lambda c: (c["lecon_id"], c["pattern_name"]))


def _feuille_vue_ensemble(ws: Worksheet, classe: dict, eleves: list[dict], date_export: str) -> None:
    """Croisement eleve x concept : maitrise (1/2/3) coloree, vide si non traverse."""
    ligne = _ecrire_entete_feuille(ws, classe, date_export)
    concepts = _concepts_de_la_classe(eleves)

    # En-tete du tableau : "Eleve" puis un concept par colonne.
    ws.cell(row=ligne, column=1, value="Eleve").font = _ENTETE_FONT
    for i, c in enumerate(concepts):
        cell = ws.cell(row=ligne, column=2 + i, value=_libelle_concept(c["pattern_name"]))
        cell.font = _ENTETE_FONT
        cell.alignment = _CENTRE
    index_col = {c["pattern_name"]: 2 + i for i, c in enumerate(concepts)}

    # Une ligne par eleve ; chaque cellule = sa maitrise sur le concept (colore).
    for r, eleve in enumerate(eleves, start=ligne + 1):
        ws.cell(row=r, column=1, value=eleve.get("prenom", ""))
        maitrise_par_concept = {
            c.get("pattern_name"): c.get("maitrise") for c in eleve.get("concepts", [])
        }
        for pattern, col in index_col.items():
            maitrise = maitrise_par_concept.get(pattern)
            if maitrise is None:
                continue
            cell = ws.cell(row=r, column=col, value=maitrise)
            cell.alignment = _CENTRE
            fill = _FILLS_MAITRISE.get(maitrise)
            if fill is not None:
                cell.fill = fill

    ws.freeze_panes = ws.cell(row=ligne + 1, column=2)  # fige en-tete + colonne eleve
    ws.column_dimensions["A"].width = 18
    for i in range(len(concepts)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16


def _feuille_detail(ws: Worksheet, classe: dict, eleves: list[dict], date_export: str) -> None:
    """Historique complet : une ligne par eleve x concept x date."""
    ligne = _ecrire_entete_feuille(ws, classe, date_export)
    entetes = ["Eleve", "Lecon", "Concept", "Maitrise", "Niveau", "Date"]
    for i, titre in enumerate(entetes):
        ws.cell(row=ligne, column=1 + i, value=titre).font = _ENTETE_FONT

    r = ligne + 1
    for eleve in eleves:
        for c in eleve.get("concepts", []):
            maitrise = c.get("maitrise")
            ws.cell(row=r, column=1, value=eleve.get("prenom", ""))
            ws.cell(row=r, column=2, value=c.get("lecon_id") or "")
            ws.cell(row=r, column=3, value=_libelle_concept(c.get("pattern_name")))
            cell_m = ws.cell(row=r, column=4, value=maitrise)
            cell_m.alignment = _CENTRE
            fill = _FILLS_MAITRISE.get(maitrise)
            if fill is not None:
                cell_m.fill = fill
            ws.cell(row=r, column=5, value=_LIBELLES_MAITRISE.get(maitrise, ""))
            ws.cell(row=r, column=6, value=c.get("date_derniere_tentative") or "")
            r += 1

    for col, largeur in zip("ABCDEF", (18, 24, 26, 10, 16, 22)):
        ws.column_dimensions[col].width = largeur


def construire_classeur(classe: dict, eleves: list[dict], date_export: str | None = None) -> Workbook:
    """Construit le classeur complet (feuilles "Vue d'ensemble" et "Detail")."""
    date_export = date_export or date.today().isoformat()
    wb = Workbook()
    ws_vue = wb.active
    ws_vue.title = "Vue d'ensemble"
    _feuille_vue_ensemble(ws_vue, classe, eleves, date_export)
    ws_detail = wb.create_sheet("Detail")
    _feuille_detail(ws_detail, classe, eleves, date_export)
    return wb
